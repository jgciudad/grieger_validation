#!/usr/bin/env python

import argparse
import sys
from importlib import import_module
from os.path import basename, join, dirname, realpath, isfile
import os
import torch
import torch.utils.data as t_data
import openpyxl
import sklearn.metrics

sys.path.insert(0, realpath(join(dirname(__file__), '..')))

from base.config_loader import ConfigLoader
from base.data.dataloader import TuebingenDataloader
from base.logger import Logger
from base.evaluation.evaluate_model import evaluate
from base.evaluation.result_logger import ResultLogger


def parse():
    parser = argparse.ArgumentParser(description='evaluate exp')
    parser.add_argument('--experiment', '-e', required=True,
                        help='name of experiment to run')
    parser.add_argument('--test_lab', '-t', required=True,
                        help="lab to test the model on: 'Antoine', 'Kornum', 'Alessandro', 'Sebastian' or 'Maiken'")
    parser.add_argument('--save_dir', '-s', required=True,
                        help="path to save model'")
    parser.add_argument('--excel_path', '-ep', required=True,
                        help="path to excel sheet to write metrics on'")
    parser.add_argument('--row', '-r', required=True,
                        help="row of frist lab to write in'")

    return parser.parse_args()


def evaluation(test_lab, r, excel_path):
    """evaluates best model in experiment on given dataset"""
    logger.fancy_log('start evaluation')
    result_logger = ResultLogger(config)

    # create dataloader for given dataset, the data should not be altered in any way
    map_loader = TuebingenDataloader(config, 'test', test_lab,
                                   data_fraction=config.DATA_FRACTION)
    dataloader = t_data.DataLoader(map_loader, batch_size=config.BATCH_SIZE_EVAL, shuffle=False, num_workers=4)

    # create empty model from model name in config and set it's state from best model in EXPERIMENT_DIR
    model = import_module('.' + config.MODEL_NAME, 'base.models').Model(config).to(config.DEVICE).eval()
    model_file = join(config.EXPERIMENT_DIR, config.MODEL_NAME + '-best.pth')
    if isfile(model_file):
        model.load_state_dict(torch.load(model_file)['state_dict'])
    else:
        raise ValueError('model_file {} does not exist'.format(model_file))
    logger.logger.info('loaded model:\n' + str(model))

    # evaluate model
    labels, _ = evaluate(config, model, dataloader)

    # log/plot results
    result_logger.log_sleep_stage_f1_scores(labels['actual'], labels['predicted'], test_lab)
    logger.logger.info('')
    result_logger.log_confusion_matrix(labels['actual'], labels['predicted'], test_lab, wo_plot=False)
    result_logger.log_transformation_matrix(labels['actual'], labels['predicted'], test_lab,
                                            wo_plot=False)
    
    # compute metrics
    recall = sklearn.metrics.recall_score(y_true=labels['actual'], y_pred=labels['predicted'], average=None)
    precision = sklearn.metrics.precision_score(y_true=labels['actual'], y_pred=labels['predicted'], average=None)
    f1score = sklearn.metrics.f1_score(y_true=labels['actual'], y_pred=labels['predicted'], average=None)
    acc = [sklearn.metrics.accuracy_score(y_true=labels['actual']==c, y_pred=labels['predicted']==c) for c in range(len(config.STAGES))]
    specificity = []
    for c in range(len(config.STAGES)):
         tn, fp, fn, tp = sklearn.metrics.confusion_matrix(y_true=labels['actual']==c, y_pred=labels['predicted']==c).ravel()
         specificity.append(tn / (tn+fp))
    bal_acc = (recall + specificity)/2
    
    wb = openpyxl.load_workbook(excel_path)   
    sheet = wb["Sheet1"]
    sheet.cell(row = r, column = 2).value = test_lab
    for idx, c in enumerate(config.STAGES):
        class_first_column = idx+3

        sheet.cell(row = r, column = 3 + 5*idx).value = recall[idx]
        sheet.cell(row = r, column = 3 + 5*idx + 1).value = precision[idx]
        sheet.cell(row = r, column = 3 + 5*idx + 2).value = f1score[idx]
        sheet.cell(row = r, column = 3 + 5*idx + 3).value = acc[idx]
        sheet.cell(row = r, column = 3 + 5*idx + 4).value = bal_acc[idx]

    wb.save(excel_path)

    logger.fancy_log('finished evaluation')


if __name__ == '__main__':
    args = parse()
    config = ConfigLoader(save_dir=os.path.join(args.save_dir, args.test_lab), experiment=args.experiment)

    logger = Logger(config)  # wrapper for logger
    logger.init_log_file(args, basename(__file__))  # create log file and log config, etc

    logger.fancy_log('evaluate best model of experiment {} on dataset {}'.format(args.experiment, args.test_lab))
    # perform evaluation
    evaluation(test_lab=args.test_lab, r=int(args.row), excel_path=args.excel_path)
